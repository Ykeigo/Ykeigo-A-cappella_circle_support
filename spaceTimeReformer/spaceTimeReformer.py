import openpyxl
import os
import openpyxl as px
print("回答の記録されたexcelファイルのファイル名を入力")

filename = input()
if not os.path.exists(filename):
    print(filename+"は存在しません")
    exit()

book = px.load_workbook(filename)
answerWorkSheet = book.active

outputWorkBook = openpyxl.Workbook()

outputWorkSheet = outputWorkBook.active

outputWorkSheet.title = "sheet1"


members = []

for row in range(0,answerWorkSheet.max_row):
    line = []
    for col in range(1,answerWorkSheet.max_column):
        line.append(answerWorkSheet.cell(row+1,col+1).value)

        #outputWorkSheet.cell(row=row+1,column=col+1,value=v)
    members.append(line)


members.sort(key=lambda x: x[0])

name = "a"
deleted = 0
m = members.copy()
for i in range(len(m)):
    if name == members[i-deleted][0]:
        members.pop(i-deleted-1)
        deleted += 1
    else:
        name = members[i-deleted][0]

members.sort(key=lambda x: x[1])

outputWorkSheet.column_dimensions[chr(ord("A") + 0)].width = 10
outputWorkSheet.column_dimensions[chr(ord("A") + 1)].width = 10
outputWorkSheet.column_dimensions[chr(ord("A") + 2)].width = 20
outputWorkSheet.column_dimensions[chr(ord("A") + 3)].width = 20
outputWorkSheet.column_dimensions[chr(ord("A") + 4)].width = 20
outputWorkSheet.column_dimensions[chr(ord("A") + 5)].width = 20
outputWorkSheet.column_dimensions[chr(ord("A") + 6)].width = 20
outputWorkSheet.column_dimensions[chr(ord("A") + 7)].width = 10
outputWorkSheet.column_dimensions[chr(ord("A") + 8)].width = 20

# あだ名、期、とかの行
for j in range(len(members[-1])):
    outputWorkSheet.cell(row=1, column=j + 1, value=members[-1][j])

#print(members)
ki = members[0][1]
reline = 0
for i in range(1,len(members)-1):
    #print(i, end = " ")
    #print(i)
    if ki != members[i][1]:
        print(ki, end=" ")
        print(members[i][1])
        # あだ名、期、とかの行
        for j in range(len(members[-1])):
            outputWorkSheet.cell(row=i+2+reline, column=j + 1, value=members[-1][j])

        ki = members[i][1]
        reline += 2

    for j in range(len(members[i])):
        if j == 7:
            outputWorkSheet.cell(row=i+1+reline,column=j+1,value=members[i][8])
        elif j == 8:
            outputWorkSheet.cell(row=i+1+reline,column=j+1,value=members[i][7])
        else:
            outputWorkSheet.cell(row=i+1+reline,column=j+1,value=members[i][j])
            if members[i][j] == "組める":
                outputWorkSheet[chr(ord("A") + j)+str(i+1+reline)].font = openpyxl.styles.fonts.Font(color='FF0000')
            elif members[i][j] == "場合によっては組める":
                outputWorkSheet[chr(ord("A") + j)+str(i+1+reline)].font = openpyxl.styles.fonts.Font(color='A00000')

print(len(members))

print(answerWorkSheet.max_column)
print(answerWorkSheet.max_row)

# 保存
outputWorkBook.save('output.xlsx')