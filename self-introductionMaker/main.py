import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
import re
from openpyxl.styles.borders import Border, Side

########読み込み#################
sourcewb = openpyxl.load_workbook("上期紹介　素材.xlsx")

#ワークシートは一個しかないはずなのでこれでok
sourcews = sourcewb.worksheets[0]

#セル一個指定の
#c1 = ws.cell(row=1, column=1)

members = []
for row in sourcews.iter_rows(min_row=2):
    addrs = []
    for cell in row:
        #print(cell.value)
        if cell.value != None:
            addrs.append(cell.value)
    #print(",".join(addrs))
    members.append(addrs)

print(len(members))
#print(members[1])
#for i in members[1]:
#    print(i)

members.pop(0)
#members = sorted(members, key=lambda x: x[2])

members.sort(key=lambda x: x[1])

name = "a"
deleted = 0
m = members.copy()
for i in range(len(m)):
    if name == members[i-deleted][1]:
        members.pop(i-deleted-1)
        deleted += 1
    else:
        name = members[i-deleted][1]

members.sort(key=lambda x: x[2])

#print(members)

#名前のよみがなはうざいので消す
kakkoPattern = re.compile('[\(（].*[\)）]')

for m in range(len(members)):
    members[m][0] = kakkoPattern.sub('',members[m][0])
"""
#名前が2行になったのでよみがなを下の行に書くという選択肢が生まれた
for m in range(len(members)):
    members[m][0] = members[m][0].replace('(', '\n(')
    members[m][0] = members[m][0].replace('（', '\n（')
"""

#samwb = openpyxl.load_workbook("上期紹介テンプレート.xlsx")
#samws = sourcewb.worksheets[0]

#print(samws.row_dimensions[1].height)
#print(samws.column_dimensions['A'].width)

##########書き込み#################

outwb = openpyxl.Workbook()

outws = outwb.worksheets[0]

#print(outws.cell(1,1).value)

#列幅指定
#列幅はどうやら1=6ピクセルらしい（憤慨）
CELLWIDTH = 1
CELLHEIGHT = 21

#鬼のマジックナンバー　今後一切Excelとjpの人紹介の仕様が変わらないことを祈る
outws.column_dimensions['A'].width = CELLWIDTH*12
outws.column_dimensions['B'].width = CELLWIDTH*9.5
outws.column_dimensions['C'].width = CELLWIDTH*28.5
outws.column_dimensions['D'].width = CELLWIDTH*2.3
outws.column_dimensions['E'].width = CELLWIDTH*12
outws.column_dimensions['F'].width = CELLWIDTH*9.5
outws.column_dimensions['G'].width = CELLWIDTH*28.5

#行幅は縦に並べて全部やるので情報を書き込みながら順にやる


mediumSide = Side(style='medium', color='000000')
hairside = Side(style='hair', color='000000')
thinside = Side(style='thin', color='000000')
#期ごとに改ページするために期が変わったら空き分をずらす
#1ページ10人(拡大率90% マジックナンバー)
pad = 0
ki = members[0][2]
for i in range(0,len(members)):
    if ki != members[i][2]:
        if (i + pad) % 10 != 0:
            pad += 10 - (i+pad)%10
        ki = members[i][2]
    homer = ((i+pad)//2)*7+1
    homec = ((i+pad)%2) *4+1

    #セルの大きさの調整は下で行う
    #文字サイズの調整も同時に行うため&開いてるとこも全部高さだけは調整したいため

    outws.cell(homer+4,homec).value = members[i][1]#サークル名
    outws.cell(homer + 4, homec).border = hairside
    outws.cell(homer+5,homec).value = members[i][0]#本名
    outws.cell(homer,homec+1).value = members[i][2]#期
    outws.cell(homer,homec+2).value = members[i][5]#パート
    outws.cell(homer+1,homec+1).value = members[i][3]#大学・学年
    outws.cell(homer+2,homec+1).value = "組んでいるバンド"
    outws.cell(homer+3,homec+1).value = members[i][4]#バンド
    outws.cell(homer+4,homec+1).value = "ひとこと"
    outws.cell(homer+5,homec+1).value = members[i][6]#ひとこと

    outws.cell(homer + 0, homec).border = Border(top=thinside, left=thinside)
    outws.cell(homer + 1, homec).border = Border(left=thinside)
    outws.cell(homer + 2, homec).border = Border(left=thinside)
    outws.cell(homer + 3, homec).border = Border(left=thinside)
    outws.cell(homer + 4, homec).border = Border(top=hairside, bottom=hairside,
                                                 left=thinside, right=hairside)
    outws.cell(homer + 5, homec).border = Border(top=hairside, bottom=thinside,
                                                 left=thinside, right=hairside)
    for j in range(7):
        for k in range(1,3):
            topside = hairside
            bottomside = hairside
            rightside = hairside
            leftside = hairside
            if j == 0:
                topside = thinside
            elif j == 5:
                bottomside = thinside
            if k == 2 or j >= 1:
                rightside = thinside
            if j != 6:
                outws.cell(homer + j, homec + k).border = Border(top=topside, bottom=bottomside,
                                                                 right=rightside,left=leftside)
    #outws.cell(homer, homec+2).border = Border(top=thinside, bottom=hairside,
    #                                            left=hairside, right=thinside)




for i in range(len(members)+pad):
    homer = ((i)//2)*7+1
    homec = ((i)%2) *4+1
    # フォント設定
    for j in range(7):
        for k in range(3):
            v = outws.cell(homer + j, homec + k).value
            if v != None and len(str(v)) > 22:
                print("too long")
                outws.cell(homer + j, homec + k).font = Font(size=8)

            outws.cell(homer + j, homec + k).alignment = Alignment(
                wrap_text=True,  # 折り返し改行
                horizontal='general',  # 水平位置
                vertical='center'  # 上下位置
            )

    # 名前はセンターがいいので調整
    outws.cell(homer + 4, homec).alignment = Alignment(
        horizontal='center',  # 水平位置
        vertical='center'
    )
    outws.cell(homer + 5, homec).alignment = Alignment(
        horizontal='center',  # 水平位置
        vertical='center'
    )
    # ひとことは2行あるのでフォントサイズを個別に調整
    if outws.cell(homer + 5, homec + 1).value != None and len(outws.cell(homer + 5, homec + 1).value) > 44:
        outws.cell(homer + 5, homec + 1).font = Font(size=8)
    else:
        outws.cell(homer + 5, homec + 1).font = Font(size=11)

    # セルの大きさ設定
    if i % 2 == 0:
        # 行高さ設定
        for j in range(7):
            outws.row_dimensions[homer + j].height = CELLHEIGHT
        # 6個目だけはひとことなので2行になってもいいように高くする
        outws.row_dimensions[homer + 5].height = CELLHEIGHT * 1.6

        # セル結合
        for j in range(1, 6):
            connectString = 'B' + str(homer + j) + ':C' + str(homer + j)
            # print(connectString)
            outws.merge_cells(connectString)
            connectString = 'F' + str(homer + j) + ':G' + str(homer + j)
            # print(connectString)
            outws.merge_cells(connectString)

#outws.cell(1, 2).font = Font(size=100)
outwb.save("formed.xlsx")
